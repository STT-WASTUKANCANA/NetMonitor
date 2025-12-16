"""
Enhanced Report Generator for NetMonitor
Generates professional PDF and Excel reports with structured chapters and formal formatting.
"""

import pandas as pd
import io
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, 
    Image, PageBreak, KeepTogether
)
from reportlab.platypus.tableofcontents import TableOfContents
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.pdfgen import canvas
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from typing import Dict, List, Any

# ============================================================================
# STYLE CONSTANTS
# ============================================================================

# PDF Colors
COLOR_PRIMARY = colors.HexColor('#1E40AF')  # Blue
COLOR_SECONDARY = colors.HexColor('#059669')  # Green
COLOR_DANGER = colors.HexColor('#DC2626')  # Red
COLOR_WARNING = colors.HexColor('#F59E0B')  # Orange
COLOR_BG_LIGHT = colors.HexColor('#F8FAFC')
COLOR_BG_MEDIUM = colors.HexColor('#E2E8F0')
COLOR_TEXT_DARK = colors.HexColor('#0F172A')

# Excel Colors
EXCEL_HEADER_BG = 'E2E8F0'
EXCEL_PRIMARY = '1E40AF'
EXCEL_SUCCESS = '059669'
EXCEL_DANGER = 'DC2626'
EXCEL_WARNING = 'F59E0B'

# ============================================================================
# PDF HELPER FUNCTIONS
# ============================================================================

class NumberedCanvas(canvas.Canvas):
    """Custom canvas for page numbering"""
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.setFont("Times-Roman", 9)
        self.drawRightString(
            A4[0] - 72, 30,
            f"Halaman {self._pageNumber} dari {page_count}"
        )
        self.drawString(
            72, 30,
            "NetMonitor - Network Monitoring Report"
        )


class ReportDocTemplate(SimpleDocTemplate):
    """Custom DocTemplate to handle Table of Contents entries"""
    def afterFlowable(self, flowable):
        "Registers TOC entries."
        if flowable.__class__.__name__ == 'Paragraph':
            text = flowable.getPlainText()
            style = flowable.style.name
            if style == 'Chapter':
                self.notify('TOCEntry', (0, text, self.page))
            elif style == 'Section':
                self.notify('TOCEntry', (1, text, self.page))
            elif style == 'Subsection':
                self.notify('TOCEntry', (2, text, self.page))

def get_pdf_styles():
    """Create and return custom PDF styles"""
    styles = getSampleStyleSheet()
    
    # Cover Title
    styles.add(ParagraphStyle(
        name='CoverTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=COLOR_PRIMARY,
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Times-Bold'
    ))
    
    # Cover Subtitle
    styles.add(ParagraphStyle(
        name='CoverSubtitle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=COLOR_TEXT_DARK,
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName='Times-Roman'
    ))
    
    # Chapter (Bab)
    styles.add(ParagraphStyle(
        name='Chapter',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=COLOR_PRIMARY,
        spaceBefore=20,
        spaceAfter=12,
        fontName='Times-Bold',
        keepWithNext=True
    ))
    
    # Section
    styles.add(ParagraphStyle(
        name='Section',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=COLOR_TEXT_DARK,
        spaceBefore=12,
        spaceAfter=8,
        fontName='Times-Bold',
        keepWithNext=True
    ))
    
    # Subsection
    styles.add(ParagraphStyle(
        name='Subsection',
        parent=styles['Heading3'],
        fontSize=11,
        textColor=COLOR_TEXT_DARK,
        spaceBefore=10,
        spaceAfter=6,
        fontName='Times-Bold',
        keepWithNext=True
    ))
    
    # Body Text
    if 'BodyText' in styles:
        s = styles['BodyText']
        s.parent = styles['Normal']
        s.fontSize = 10
        s.textColor = COLOR_TEXT_DARK
        s.fontName = 'Times-Roman'
        s.alignment = TA_JUSTIFY
        s.spaceAfter = 6
    else:
        styles.add(ParagraphStyle(
            name='BodyText',
            parent=styles['Normal'],
            fontSize=10,
            textColor=COLOR_TEXT_DARK,
            fontName='Times-Roman',
            alignment=TA_JUSTIFY,
            spaceAfter=6
        ))
    
    # Caption
    if 'Caption' in styles:
        s = styles['Caption']
        s.parent = styles['Normal']
        s.fontSize = 9
        s.textColor = colors.grey
        s.fontName = 'Times-Italic'
        s.alignment = TA_CENTER
    else:
        styles.add(ParagraphStyle(
            name='Caption',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            fontName='Times-Italic',
            alignment=TA_CENTER
        ))
    
    # Footer
    styles.add(ParagraphStyle(
        name='Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        fontName='Times-Italic',
        alignment=TA_CENTER
    ))
    
    return styles


def create_cover_page(data: dict, styles) -> List:
    """Create professional cover page with optional campus logo"""
    from pathlib import Path
    import os
    
    story = []
    
    # Add spacing from top
    story.append(Spacer(1, 1*inch))
    
    # Try to add campus logo if available
    logo_path = Path(__file__).parent.parent / 'assets' / 'logo_kampus.png'
    if logo_path.exists():
        try:
            logo = Image(str(logo_path), width=1.5*inch, height=1.5*inch)
            logo.hAlign = 'CENTER'
            story.append(logo)
            story.append(Spacer(1, 0.3*inch))
        except Exception:
            # If logo fails to load, just skip it
            pass
    
    # Title - more friendly/casual
    story.append(Paragraph("NetMonitor", styles['CoverTitle']))
    story.append(Paragraph("Laporan Monitoring Jaringan Kampus", styles['CoverSubtitle']))
    
    story.append(Spacer(1, 0.5*inch))
    
    # Period info box
    period_data = [
        ['Periode Monitoring', data.get('period', 'N/A')],
        ['Tanggal Mulai', data.get('period_start', 'N/A')],
        ['Tanggal Akhir', data.get('period_end', 'N/A')],
        ['Tanggal Generate', data.get('generated_at', 'N/A')]
    ]
    
    t = Table(period_data, colWidths=[2.5*inch, 2.5*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), COLOR_BG_MEDIUM),
        ('BACKGROUND', (1, 0), (1, -1), colors.white),
        ('TEXTCOLOR', (0, 0), (-1, -1), COLOR_TEXT_DARK),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Times-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Times-Roman'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 1, COLOR_PRIMARY),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    story.append(t)
    
    story.append(Spacer(1, 1*inch))
    
    # Document info - shorter and more casual
    story.append(Paragraph(
        "Laporan ini memberikan gambaran lengkap kondisi jaringan kampus berdasarkan monitoring otomatis NetMonitor System.",
        styles['BodyText']
    ))
    
    story.append(Spacer(1, 0.5*inch))
    story.append(Paragraph(
        "<i>Generated by NetMonitor System</i>",
        styles['Footer']
    ))
    
    story.append(PageBreak())
    
    return story


def create_toc(styles) -> TableOfContents:
    """Create Table of Contents"""
    toc = TableOfContents()
    toc.levelStyles = [
        ParagraphStyle(
            name='TOCHeading1',
            fontSize=12,
            fontName='Times-Bold',
            leftIndent=20,
            spaceBefore=5,
            spaceAfter=5
        ),
        ParagraphStyle(
            name='TOCHeading2',
            fontSize=10,
            fontName='Times-Roman',
            leftIndent=40,
            spaceBefore=2,
            spaceAfter=2
        ),
    ]
    return toc


def create_chart_image(daily_stats: List[dict], metric_key: str, title: str, 
                       ylabel: str, color: str, kind: str = 'line') -> io.BytesIO:
    """Create high-quality chart image"""
    if not daily_stats:
        return None
        
    dates = [d['date'] for d in daily_stats]
    values = [d.get(metric_key, 0) for d in daily_stats]
    
    plt.figure(figsize=(7, 3.5))
    plt.style.use('seaborn-v0_8-darkgrid')
    
    # Adaptive X-axis formatting
    try:
        if len(dates) > 0:
            # Detect format based on string length or content
            sample_date = str(dates[0])
            is_hourly = len(sample_date) > 10 # "YYYY-MM-DD HH:MM"
            
            # Create actual datetime objects for better plotting
            if is_hourly:
                fmt = '%Y-%m-%d %H:%M:%S' if len(sample_date) > 16 else '%Y-%m-%d %H:%M'
                dt_objects = [datetime.strptime(str(d), fmt) for d in dates]
                
                # Plot using dates
                if kind == 'bar':
                    plt.bar(dt_objects, values, color=color, alpha=0.8, edgecolor='white', linewidth=0.7, width=0.03 if len(dates) > 50 else 0.1)
                else:
                    plt.plot(dt_objects, values, marker='o' if len(dates) < 50 else None, 
                            linestyle='-', color=color, linewidth=2.5, 
                            markersize=6, markerfacecolor='white', 
                            markeredgewidth=2, markeredgecolor=color)
                    plt.fill_between(dt_objects, values, alpha=0.2, color=color)
                
                # Format Axis
                ax = plt.gca()
                # If range < 2 days, show Hours
                if (dt_objects[-1] - dt_objects[0]).total_seconds() < 172800: # 48h
                     ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
                     plt.xlabel('Jam' if kind!='bar' else '', fontsize=10, fontweight='bold')
                else: 
                     # Show Date
                     ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
                     plt.xlabel('Tanggal', fontsize=10, fontweight='bold')
                     
                plt.gcf().autofmt_xdate()
            else:
                 # Standard Daily YYYY-MM-DD
                 if kind == 'bar':
                     plt.bar(dates, values, color=color, alpha=0.8, edgecolor='white', linewidth=0.7)
                 else:
                     plt.plot(dates, values, marker='o', linestyle='-', color=color, 
                             linewidth=2.5, markersize=6, markerfacecolor='white', 
                             markeredgewidth=2, markeredgecolor=color)
                     plt.fill_between(dates, values, alpha=0.2, color=color)
                 plt.xlabel('Tanggal', fontsize=10, fontweight='bold')
    except Exception as e:
        # Fallback to simple string plotting
        pass

        
    plt.title(title, fontsize=12, fontweight='bold', pad=15)
    plt.ylabel(ylabel, fontsize=10, fontweight='bold')
    plt.xticks(rotation=45, ha='right', fontsize=9)
    plt.yticks(fontsize=9)
    plt.grid(True, linestyle='--', alpha=0.3)
    plt.tight_layout()
    
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close()
    buf.seek(0)
    return buf


def create_pie_chart_image(data_dict: dict, title: str, colors_list: List[str] = None) -> io.BytesIO:
    """Create high-quality pie chart for severity distribution"""
    if not data_dict or sum(data_dict.values()) == 0:
        return None
    
    # Filter out zero values
    filtered_data = {k: v for k, v in data_dict.items() if v > 0}
    if not filtered_data:
        return None
    
    labels = list(filtered_data.keys())
    values = list(filtered_data.values())
    
    # Default colors if not provided
    if not colors_list:
        colors_list = ['#EF4444', '#F97316', '#F59E0B', '#10B981']
    
    plt.figure(figsize=(6, 4))
    plt.pie(values, labels=labels, autopct='%1.1f%%', startangle=90, 
            colors=colors_list[:len(values)], textprops={'fontsize': 10, 'fontweight': 'bold'})
    plt.title(title, fontsize=12, fontweight='bold', pad=15)
    plt.axis('equal')
    plt.tight_layout()
    
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close()
    buf.seek(0)
    return buf


def create_summary_metrics_table(summary: dict, alert_stats: dict) -> Table:
    """Create executive summary metrics table"""
    data = [
        ['METRIK JARINGAN', 'NILAI', 'METRIK ALERT', 'NILAI'],
        ['Total Perangkat', str(summary.get('total_devices', 0)),
         'Total Alert', str(alert_stats.get('total_alerts', 0))],
        ['Perangkat Online', str(summary.get('online_count', 0)),
         'Alert Aktif', str(alert_stats.get('active_count', 0))],
        ['Perangkat Offline', str(summary.get('offline_count', 0)),
         'Alert Critical', str(alert_stats.get('critical_count', 0))],
        ['Uptime (%)', f"{summary.get('uptime_percentage', 0)}%",
         'Alert Resolved', str(alert_stats.get('resolved_count', 0))],
        ['Rata-rata Latency', f"{summary.get('avg_response_time', 0)} ms",
         'MTTR (menit)', f"{alert_stats.get('avg_resolution_time_minutes', 0)} min"]
    ]
    
    t = Table(data, colWidths=[1.8*inch, 1.2*inch, 1.8*inch, 1.2*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), COLOR_PRIMARY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
        ('FONTNAME', (0, 1), (0, -1), 'Times-Bold'),
        ('FONTNAME', (2, 1), (2, -1), 'Times-Bold'),
        ('FONTNAME', (1, 1), (-1, -1), 'Times-Roman'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, COLOR_BG_LIGHT]),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    return t


def generate_recommendations(summary: dict, alert_stats: dict, top_issues: List[dict], daily_stats: List[dict] = None) -> List[str]:
    """Generate automated recommendations based on data - casual but informative format"""
    recommendations = []
    
    # Check availability
    uptime = summary.get('uptime_percentage', 0)
    avg_latency = summary.get('avg_response_time', 0)
    critical_count = alert_stats.get('critical_count', 0)
    mttr = alert_stats.get('avg_resolution_time_minutes', 0)
    
    # Priority 1: Critical alerts
    if critical_count > 5:
        recommendations.append(
            f"⚠️ Ada masalah serius - {critical_count} critical alerts perlu ditangani segera"
        )
    elif critical_count > 0:
        recommendations.append(
            f"🔔 {critical_count} alert critical - prioritaskan penanganan untuk mencegah dampak lebih luas"
        )
    
    # Priority 2: Availability issue
    if uptime < 90:
        recommendations.append(
            f"🔴 Perlu pengecekan segera - Availability {uptime}% di bawah standar minimum 90%"
        )
    elif uptime < 95:
        recommendations.append(
            f"🟡 Availability {uptime}% perlu ditingkatkan - investigasi perangkat yang sering down"
        )
    elif uptime >= 99:
        recommendations.append(
            f"✅ Uptime sangat baik ({uptime}%) - pertahankan standar monitoring"
        )
    
    # Check performance
    if avg_latency > 100:
        recommendations.append(
            f"⚡ Latency tinggi ({avg_latency} ms) - periksa konfigurasi jaringan dan bandwidth"
        )
    
    # Check MTTR
    if mttr > 60:
        recommendations.append(
            f"⏱️ MTTR {mttr} menit terlalu lama - tingkatkan proses incident response"
        )
    
    # Check trend (if daily stats available)
    if daily_stats and len(daily_stats) >= 2:
        try:
            recent_uptime = sum([d.get('uptime', 0) for d in daily_stats[-3:]]) / min(3, len(daily_stats))
            older_uptime = sum([d.get('uptime', 0) for d in daily_stats[:3]]) / min(3, len(daily_stats))
            if recent_uptime < older_uptime - 25:  # 25% drop
                recommendations.append(
                    "📉 Trend negatif terdeteksi - downtime meningkat signifikan dalam periode ini"
                )
        except:
            pass  # Skip trend analysis if data insufficient
    
    # Check top issues
    if len(top_issues) >= 5:
        recommendations.append(
            f"🔧 {len(top_issues)} perangkat dengan issue tinggi - lakukan audit mendalam"
        )
    
    # Perfect case
    if uptime == 100 and critical_count == 0:
        recommendations.append(
            "🎉 Semua sistem berjalan normal - kondisi jaringan excellent!"
        )
    
    # Ensure at least one recommendation
    if not recommendations:
        recommendations.append(
            "📊 Kondisi jaringan stabil - lanjutkan monitoring rutin"
        )
    
    return recommendations


# ============================================================================
# PDF GENERATION MAIN FUNCTION
# ============================================================================

def generate_pdf_report(data: dict) -> bytes:
    """
    Generate professional PDF report with structured chapters.
    Returns bytes of the PDF file.
    """
    buffer = io.BytesIO()
    doc = ReportDocTemplate(
        buffer, 
        pagesize=A4, 
        rightMargin=72, 
        leftMargin=72, 
        topMargin=72, 
        bottomMargin=50
    )
    
    story = []
    styles = get_pdf_styles()
    
    # Extract data
    summary = data.get('summary', {})
    alert_stats = data.get('alert_stats', {})
    daily_stats = data.get('daily_stats', [])
    top_issues = data.get('top_issues', [])
    recent_alerts = data.get('recent_alerts', [])
    
    # ========================================================================
    # COVER PAGE
    # ========================================================================
    story.extend(create_cover_page(data, styles))
    
    # ========================================================================
    # TABLE OF CONTENTS
    # ========================================================================
    story.append(Paragraph("DAFTAR ISI", styles['Chapter']))
    toc = create_toc(styles)
    story.append(toc)
    story.append(PageBreak())
    
    # ========================================================================
    # OVERVIEW / RINGKASAN (Skip formal Pendahuluan)
    # ========================================================================
    story.append(Paragraph("RINGKASAN MONITORING", styles['Chapter']))
    
    story.append(Paragraph("Periode Monitoring", styles['Section']))
    period_text = f"Data monitoring dari <b>{data.get('period_start', 'N/A')}</b> sampai <b>{data.get('period_end', 'N/A')}</b>."
    story.append(Paragraph(period_text, styles['BodyText']))
    
    story.append(Spacer(1, 12))
    story.append(Paragraph("Metrik Utama", styles['Section']))
    
    # Summary metrics table
    story.append(create_summary_metrics_table(summary, alert_stats))
    story.append(Spacer(1, 12))
    
    # Quick insights - bullet points
    story.append(Paragraph("Key Points:", styles['Section']))
    
    uptime = summary.get('uptime_percentage', 0)
    status = "sangat baik" if uptime >= 99 else "baik" if uptime >= 95 else "perlu perhatian"
    
    insights = [
        f"• Status jaringan: <b>{status}</b> dengan uptime {uptime}%",
        f"• Total {summary.get('total_devices', 0)} perangkat dimonitor, {summary.get('online_count', 0)} online",
        f"• Rata-rata response time: {summary.get('avg_response_time', 0)} ms",
        f"• {alert_stats.get('total_alerts', 0)} alert tercatat, {alert_stats.get('critical_count', 0)} critical"
    ]
    
    for insight in insights:
        story.append(Paragraph(insight, styles['BodyText']))
    
    story.append(PageBreak())
    
    # ========================================================================
    # BAB II: RINGKASAN EKSEKUTIF
    # ========================================================================
    story.append(Paragraph("BAB II", styles['Chapter']))
    story.append(Paragraph("RINGKASAN EKSEKUTIF", styles['Chapter']))
    
    story.append(Paragraph("2.1 Overview Kondisi Jaringan", styles['Section']))
    
    # Summary metrics table
    story.append(create_summary_metrics_table(summary, alert_stats))
    story.append(Spacer(1, 12))
    
    story.append(Paragraph("2.2 Kesimpulan Utama", styles['Section']))
    
    # Generate dynamic conclusion
    uptime = summary.get('uptime_percentage', 0)
    status = "sangat baik" if uptime >= 99 else "baik" if uptime >= 95 else "perlu perhatian"
    
    conclusion = f"Berdasarkan data monitoring dalam periode ini, kondisi jaringan berada dalam status <b>{status}</b> " \
                 f"dengan uptime sebesar <b>{uptime}%</b>. "
    
    if alert_stats.get('critical_count', 0) > 0:
        conclusion += f"Terdapat <b>{alert_stats.get('critical_count')} alert critical</b> yang memerlukan perhatian segera. "
    
    conclusion += f"Rata-rata response time jaringan adalah <b>{summary.get('avg_response_time', 0)} ms</b>."
    
    story.append(Paragraph(conclusion, styles['BodyText']))
    
    story.append(PageBreak())
    
    # ========================================================================
    # BAB III: ANALISIS PERFORMA JARINGAN
    # ========================================================================
    story.append(Paragraph("BAB III", styles['Chapter']))
    story.append(Paragraph("ANALISIS PERFORMA JARINGAN", styles['Chapter']))
    
    story.append(Paragraph("3.1 Trend Response Time", styles['Section']))
    
    if daily_stats:
        chart_latency = create_chart_image(
            daily_stats, 'avg_latency', 
            'Trend Response Time Harian', 
            'Latency (ms)', '#1E40AF', kind='line'
        )
        if chart_latency:
            img = Image(chart_latency, width=6*inch, height=3*inch)
            story.append(img)
            story.append(Paragraph(
                "Grafik 3.1: Trend response time menunjukkan rata-rata latency jaringan per hari dalam periode monitoring.",
                styles['Caption']
            ))
    else:
        story.append(Paragraph("Data trend response time tidak tersedia untuk periode ini.", styles['BodyText']))
    
    story.append(Spacer(1, 12))
    story.append(Paragraph("3.2 Trend Uptime Jaringan", styles['Section']))
    
    if daily_stats:
        chart_uptime = create_chart_image(
            daily_stats, 'uptime', 
            'Trend Uptime Harian', 
            'Uptime (%)', '#059669', kind='line'
        )
        if chart_uptime:
            img = Image(chart_uptime, width=6*inch, height=3*inch)
            story.append(img)
            story.append(Paragraph(
                "Grafik 3.2: Trend uptime menunjukkan persentase ketersediaan jaringan per hari.",
                styles['Caption']
            ))
    
    story.append(Spacer(1, 12))
    story.append(Paragraph("3.3 Status Perangkat", styles['Section']))
    
    device_summary = f"Total perangkat yang dimonitor: <b>{summary.get('total_devices', 0)}</b><br/>" \
                     f"Perangkat Online: <b>{summary.get('online_count', 0)}</b><br/>" \
                     f"Perangkat Offline: <b>{summary.get('offline_count', 0)}</b><br/>" \
                     f"Perangkat Unknown: <b>{summary.get('unknown_count', 0)}</b>"
    
    story.append(Paragraph(device_summary, styles['BodyText']))
    
    story.append(PageBreak())
    
    # ========================================================================
    # ANALISIS ALERT DAN INSIDEN
    # ========================================================================
    story.append(Paragraph("ANALISIS ALERT", styles['Chapter']))
    
    story.append(Paragraph("Distribusi Severity Alert", styles['Section']))
    
    # Create pie chart for severity distribution
    severity_data = {
        'Critical': alert_stats.get('critical_count', 0),
        'High': alert_stats.get('high_count', 0),
        'Medium': alert_stats.get('medium_count', 0),
        'Low': alert_stats.get('low_count', 0)
    }
    
    severity_colors = ['#EF4444', '#F97316', '#F59E0B', '#10B981']
    pie_chart = create_pie_chart_image(severity_data, 'Distribusi Alert by Severity', severity_colors)
    
    if pie_chart:
        img = Image(pie_chart, width=5*inch, height=3.5*inch)
        story.append(img)
        story.append(Paragraph(
            f"Total {alert_stats.get('total_alerts', 0)} alert: Critical ({alert_stats.get('critical_count', 0)}), "\
            f"High ({alert_stats.get('high_count', 0)}), Medium ({alert_stats.get('medium_count', 0)}), "\
            f"Low ({alert_stats.get('low_count', 0)})",
            styles['Caption']
        ))
    
    story.append(Spacer(1, 12))
    story.append(Paragraph("Trend Volume Alert", styles['Section']))
    
    if daily_stats:
        chart_alerts = create_chart_image(
            daily_stats, 'alert_count', 
            'Volume Alert per Hari', 
            'Jumlah Alert', '#F59E0B', kind='bar'
        )
        if chart_alerts:
            img = Image(chart_alerts, width=6*inch, height=3*inch)
            story.append(img)
            story.append(Paragraph(
                "Trend jumlah alert yang terjadi setiap hari.",
                styles['Caption']
            ))
    
    story.append(Spacer(1, 12))
    story.append(Paragraph("4.3 Perangkat dengan Issue Tertinggi", styles['Section']))
    
    if top_issues:
        issue_data = [['Nama Perangkat', 'IP Address', 'Jumlah Alert', 'Est. Downtime (menit)']]
        for issue in top_issues[:10]:  # Top 10
            issue_data.append([
                issue.get('device_name', 'N/A'),
                issue.get('ip_address', 'N/A'),
                str(issue.get('alert_count', 0)),
                f"{issue.get('downtime_minutes', 0):.1f}"
            ])
        
        t_issues = Table(issue_data, colWidths=[2*inch, 1.5*inch, 1.2*inch, 1.3*inch])
        t_issues.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_DANGER),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Times-Roman'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, COLOR_BG_LIGHT]),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(t_issues)
        story.append(Paragraph(
            "Tabel 4.1: Daftar perangkat dengan jumlah alert dan estimasi downtime tertinggi.",
            styles['Caption']
        ))
    else:
        story.append(Paragraph("Tidak ada perangkat dengan issue signifikan dalam periode ini.", styles['BodyText']))
    
    story.append(Spacer(1, 12))
    story.append(Paragraph("4.4 Alert Terbaru", styles['Section']))
    
    if recent_alerts:
        alert_data = [['Waktu', 'Severity', 'Pesan', 'Status']]
        for alert in recent_alerts[:15]:
            alert_data.append([
                alert.get('created_at', '')[:16].replace('T', ' '),
                alert.get('severity', '').upper(),
                Paragraph(alert.get('message', ''), styles['BodyText']),
                alert.get('status', '').title()
            ])
        
        t_alerts = Table(alert_data, colWidths=[1.3*inch, 0.9*inch, 2.8*inch, 1*inch])
        t_alerts.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_WARNING),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Times-Roman'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, COLOR_BG_LIGHT]),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(t_alerts)
        story.append(Paragraph(
            "Tabel 4.2: Daftar 15 alert terbaru yang tercatat dalam sistem.",
            styles['Caption']
        ))
    else:
        story.append(Paragraph("Tidak ada alert yang tercatat dalam periode ini.", styles['BodyText']))
    
    story.append(PageBreak())
    
    # ========================================================================
    # REKOMENDASI
    # ========================================================================
    story.append(Paragraph("REKOMENDASI", styles['Chapter']))
    
    story.append(Paragraph("Action Items", styles['Section']))
    story.append(Paragraph(
        "Berdasarkan analisis data monitoring, berikut adalah rekomendasi untuk peningkatan jaringan:",
        styles['BodyText']
    ))
    
    story.append(Spacer(1, 10))
    
    # Generate recommendations with trend data
    recommendations = generate_recommendations(summary, alert_stats, top_issues, daily_stats)
    
    for i, rec in enumerate(recommendations, 1):
        story.append(Paragraph(rec, styles['BodyText']))
        story.append(Spacer(1, 6))
    
    story.append(Spacer(1, 30))
    story.append(Paragraph(
        f"--- End of Report ---<br/>Generated automatically by NetMonitor System on {data.get('generated_at', 'N/A')}",
        styles['Footer']
    ))
    
    # Build PDF with numbered pages
    # Build PDF
    doc.multiBuild(story, canvasmaker=NumberedCanvas)
    
    buffer.seek(0)
    return buffer.getvalue()


# ============================================================================
# EXCEL GENERATION
# ============================================================================

def generate_excel_report(data: dict) -> bytes:
    """
    Generate professional Excel report with multiple formatted sheets.
    Returns bytes of the Excel file.
    """
    output = io.BytesIO()
    
    # Extract data
    summary = data.get('summary', {})
    alert_stats = data.get('alert_stats', {})
    daily_stats = data.get('daily_stats', [])
    top_issues = data.get('top_issues', [])
    
    # New data fields
    all_devices_stats = data.get('all_devices_stats', [])
    all_alerts = data.get('all_alerts', [])
    raw_data = data.get('raw_data', [])
    
    # Fallback to recent_alerts if all_alerts not available (for backward compatibility)
    if not all_alerts:
        all_alerts = data.get('recent_alerts', [])

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # ====================================================================
        # SHEET 1: DASHBOARD
        # ====================================================================
        dashboard_data = {
            'Item': [
                'LAPORAN MONITORING JARINGAN',
                'NetMonitor System',
                '',
                'INFORMASI LAPORAN',
                'Periode',
                'Tanggal Mulai',
                'Tanggal Akhir',
                'Tanggal Generate',
                '',
                'RINGKASAN EKSEKUTIF',
                'Total Perangkat',
                'Perangkat Online',
                'Perangkat Offline',
                'Uptime (%)',
                'Rata-rata Latency (ms)',
                '',
                'STATISTIK ALERT',
                'Total Alert',
                'Alert Aktif',
                'Alert Critical',
                'Alert Resolved',
                'MTTR (menit)'
            ],
            'Value': [
                '',
                '',
                '',
                '',
                data.get('period', 'N/A'),
                data.get('period_start', 'N/A'),
                data.get('period_end', 'N/A'),
                data.get('generated_at', 'N/A'),
                '',
                '',
                summary.get('total_devices', 0),
                summary.get('online_count', 0),
                summary.get('offline_count', 0),
                f"{summary.get('uptime_percentage', 0)}%",
                summary.get('avg_response_time', 0),
                '',
                '',
                alert_stats.get('total_alerts', 0),
                alert_stats.get('active_count', 0),
                alert_stats.get('critical_count', 0),
                alert_stats.get('resolved_count', 0),
                alert_stats.get('avg_resolution_time_minutes', 0)
            ]
        }
        
        df_dash = pd.DataFrame(dashboard_data)
        df_dash.to_excel(writer, sheet_name='Dashboard', index=False, startrow=1)
        
        ws_dash = writer.sheets['Dashboard']
        
        # Format dashboard
        ws_dash['A2'].font = Font(name='Calibri', size=18, bold=True, color=EXCEL_PRIMARY)
        ws_dash['A3'].font = Font(name='Calibri', size=12, italic=True)
        
        for row in [5, 11, 18]:
            ws_dash[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            ws_dash[f'A{row}'].fill = PatternFill(start_color=EXCEL_PRIMARY, end_color=EXCEL_PRIMARY, fill_type='solid')
            
        ws_dash.column_dimensions['A'].width = 30
        ws_dash.column_dimensions['B'].width = 25
        
        # ====================================================================
        # SHEET 2: DEVICE STATUS
        # ====================================================================
        if all_devices_stats:
            dev_rows = []
            for d in all_devices_stats:
                dev_rows.append({
                    'Device Name': d.get('name'),
                    'IP Address': d.get('ip_address'),
                    'Type': d.get('type'),
                    'Location': d.get('location'),
                    'Uptime (%)': d.get('uptime_percentage'),
                    'Downtime (Hours)': d.get('downtime_hours'),
                    'Total Alerts': d.get('alert_count'),
                    'Status': d.get('status').upper()
                })
            
            # Sort by uptime ascending (lowest first as requested)
            dev_rows.sort(key=lambda x: x['Uptime (%)'])
            
            df_dev = pd.DataFrame(dev_rows)
            df_dev.to_excel(writer, sheet_name='Device Status', index=False)
            
            ws_dev = writer.sheets['Device Status']
            
            # Header format
            for cell in ws_dev[1]:
                cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=EXCEL_PRIMARY, end_color=EXCEL_PRIMARY, fill_type='solid')
            
            # Data format & Conditional Formatting
            for row in range(2, len(dev_rows) + 2):
                uptime_val = ws_dev[f'E{row}'].value
                status_cell = ws_dev[f'H{row}']
                
                # Uptime conditional formatting
                if isinstance(uptime_val, (int, float)):
                    if uptime_val >= 95:
                        ws_dev[f'E{row}'].font = Font(color=EXCEL_SUCCESS, bold=True)
                    elif uptime_val >= 90:
                        ws_dev[f'E{row}'].font = Font(color=EXCEL_WARNING, bold=True)
                    else:
                        ws_dev[f'E{row}'].font = Font(color=EXCEL_DANGER, bold=True)
                
                # Status formatting
                if status_cell.value == 'DOWN':
                    status_cell.fill = PatternFill(start_color=EXCEL_DANGER, end_color=EXCEL_DANGER, fill_type='solid')
                    status_cell.font = Font(color='FFFFFF', bold=True)
                elif status_cell.value == 'UP':
                    status_cell.font = Font(color=EXCEL_SUCCESS, bold=True)

            # Auto-filter
            ws_dev.auto_filter.ref = ws_dev.dimensions
            
            # Widths
            ws_dev.column_dimensions['A'].width = 25
            ws_dev.column_dimensions['B'].width = 15
            ws_dev.column_dimensions['E'].width = 12

        # ====================================================================
        # SHEET 3: TREND SUMMARY (DAILY/HOURLY)
        # ====================================================================
        if daily_stats:
            df_daily = pd.DataFrame(daily_stats)
            df_daily.rename(columns={
                'date': 'Timestamp',
                'uptime': 'Uptime (%)',
                'avg_latency': 'Avg Latency (ms)',
                'alert_count': 'Total Alerts'
            }, inplace=True)
            
            # Select columns
            cols = ['Timestamp', 'Uptime (%)', 'Avg Latency (ms)', 'Total Alerts']
            # Filter existing columns
            cols = [c for c in cols if c in df_daily.columns]
            df_daily = df_daily[cols]
            
            df_daily.to_excel(writer, sheet_name='Trend Summary', index=False)
            
            ws_daily = writer.sheets['Trend Summary']
            
            # Format header
            for cell in ws_daily[1]:
                cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=EXCEL_PRIMARY, end_color=EXCEL_PRIMARY, fill_type='solid')
                
            # Column widths
            ws_daily.column_dimensions['A'].width = 20
            
            # Add sparkline-like chart (Line Chart for Uptime)
            rec_len = len(df_daily)
            if rec_len > 1:
                chart = LineChart()
                chart.title = "Uptime Trend"
                chart.style = 12
                chart.y_axis.title = "Uptime %"
                chart.x_axis.title = "Period"
                
                data_ref = Reference(ws_daily, min_col=2, min_row=1, max_row=rec_len+1)
                cats_ref = Reference(ws_daily, min_col=1, min_row=2, max_row=rec_len+1)
                
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                
                ws_daily.add_chart(chart, "F2")

        # ====================================================================
        # SHEET 4: ALERT DETAILS
        # ====================================================================
        if all_alerts:
            alert_rows = []
            for a in all_alerts:
                # Need to handle device info which is nested or flattened
                d_name = a.get('device', {}).get('name', 'Unknown') if isinstance(a.get('device'), dict) else 'Unknown'
                d_ip = a.get('device', {}).get('ip_address', '') if isinstance(a.get('device'), dict) else ''
                
                alert_rows.append({
                    'Timestamp': a.get('created_at', '').replace('T', ' ')[:19],
                    'Device': f"{d_name} ({d_ip})",
                    'Severity': a.get('severity', '').upper(),
                    'Message': a.get('message'),
                    'Status': a.get('status', '').upper(),
                    'Resolved At': a.get('resolved_at', '').replace('T', ' ')[:19] if a.get('resolved_at') else '-'
                })
            
            df_alerts = pd.DataFrame(alert_rows)
            df_alerts.to_excel(writer, sheet_name='Alert Details', index=False)
            ws_alerts = writer.sheets['Alert Details']
            
            # Header
            for cell in ws_alerts[1]:
                cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=EXCEL_WARNING, end_color=EXCEL_WARNING, fill_type='solid')

            # Conditional Formatting for Severity
            for row in range(2, len(alert_rows) + 2):
                sev_cell = ws_alerts[f'C{row}'] 
                if sev_cell.value == 'CRITICAL':
                    sev_cell.fill = PatternFill(start_color=EXCEL_DANGER, end_color=EXCEL_DANGER, fill_type='solid')
                    sev_cell.font = Font(color='FFFFFF', bold=True)
                elif sev_cell.value == 'WARNING': # If 'high' maps to warning logic or strictly enum?
                    pass 
                elif sev_cell.value == 'HIGH':
                     sev_cell.font = Font(color=EXCEL_WARNING, bold=True)

            ws_alerts.column_dimensions['A'].width = 20
            ws_alerts.column_dimensions['B'].width = 30
            ws_alerts.column_dimensions['D'].width = 50

        # ====================================================================
        # SHEET 5: TOP PROBLEMS
        # ====================================================================
        if top_issues:
            prob_rows = []
            for i in top_issues:
                prob_rows.append({
                    'Device Name': i.get('device_name'),
                    'IP Address': i.get('ip_address'),
                    'Total Alerts': i.get('alert_count'),
                    'Est. Downtime (Min)': i.get('downtime_minutes')
                })
            
            df_probs = pd.DataFrame(prob_rows)
            df_probs.to_excel(writer, sheet_name='Top Problems', index=False)
            ws_probs = writer.sheets['Top Problems']
            
            for cell in ws_probs[1]:
                cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=EXCEL_DANGER, end_color=EXCEL_DANGER, fill_type='solid')

            # Bar Chart
            if len(prob_rows) > 0:
                chart = BarChart()
                chart.title = "Top Alert Sources"
                data_ref = Reference(ws_probs, min_col=3, min_row=1, max_row=min(len(prob_rows)+1, 16)) # Alerts column
                cats_ref = Reference(ws_probs, min_col=1, min_row=2, max_row=min(len(prob_rows)+1, 16)) # Names
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                ws_probs.add_chart(chart, "F2")

        # ====================================================================
        # SHEET 6: RAW DATA
        # ====================================================================
        if raw_data:
            # Flatten raw data if needed or just dump
            # DeviceLogResponse has id, device_id, status, response_time, packet_loss, checked_at
            raw_rows = []
            for r in raw_data:
                raw_rows.append({
                    'Log ID': r.get('id'),
                    'Device ID': r.get('device_id'),
                    'Status': r.get('status'),
                    'Response Time (ms)': r.get('response_time'),
                    'Packet Loss (%)': r.get('packet_loss'),
                    'Timestamp': r.get('checked_at', '').replace('T', ' ')[:19]
                })
                
            df_raw = pd.DataFrame(raw_rows)
            df_raw.to_excel(writer, sheet_name='Raw Data', index=False)
    
    return output.getvalue()
